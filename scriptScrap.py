''' 
Dependências, só instalar na primeira vez
!pip install selenium
!pip install bs4
!pip install openpyxl
'''

import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from bs4 import BeautifulSoup
import time
import re
import threading  # Importe a biblioteca threading

def initialize_driver():
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])
    return webdriver.Chrome(options=chrome_options)

def get_page_content(driver, url):
    driver.get(url)
    driver.implicitly_wait(10)
    time.sleep(2)
    return driver.page_source

def extract_process_info(content):
    soup = BeautifulSoup(content, 'html.parser')
    titulo_elements = soup.find_all('div', class_='v-toolbar__title')
    for titulo_element in titulo_elements:
        titulo_text = titulo_element.get_text(strip=True)
    titulo = titulo_element.get_text()
    linhas = [linha.strip() for linha in titulo.split('\n') if linha.strip()]
    if len(linhas) >= 2:
        tipo_documento = linhas[0]
        numero_processo = linhas[1]
        return tipo_documento, numero_processo
    return None, None

def extract_textarea_value(driver, css_selector):
    textarea_element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, css_selector)))
    return textarea_element.get_attribute('value')

def is_valid_url(url):
    pattern = r'https://protocolo\.ufes\.br/#/documentos/\d+/'
    return re.match(pattern, url) is not None

def process_data():
    url = url_entry.get()

    if not url:
        messagebox.showerror("Erro", "Por favor, insira a URL do processo.")
        return
    elif not (is_valid_url(url)):
        messagebox.showerror("Erro", "A url está errada.")
        return
    
    # Atualize a mensagem de carregamento na interface
    loading_label.config(text="Carregando dados...")

    try:
        driver = initialize_driver()
        content = get_page_content(driver, url)
        tipo_documento, numero_processo = extract_process_info(content)

        interessado_text = extract_textarea_value(driver, "textarea[aria-label='Interessado']")
        resumo_text = extract_textarea_value(driver, "textarea[aria-label='Resumo']")

        driver.quit()

        # Abre a página e tramitações do processo
        driver = initialize_driver()
        url_tramitacoes = url + 'tramitacoes'
        content = get_page_content(driver, url_tramitacoes)

        # Clina no elemento da página para que o conteúdo do processo apareça
        tr_element = driver.find_element(By.CLASS_NAME, 'pointer')
        tr_element.click()

        despacho_text = extract_textarea_value(driver, "textarea[aria-label='Despacho']")

        # Fecha a página
        driver.quit()

        despacho = despacho_text.split('\n\n')[0]

        dados = [
            {
                "Tipo de documento": tipo_documento,
                "Número do processo": numero_processo,
                "Interessado": interessado_text,
                "Resumo": resumo_text,
                "Despacho": despacho
            }
        ]

        file_name = "processos-lepisma.xlsx"

        try:
            workbook = load_workbook(file_name)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Tipo de documento", "Número do processo", "Interessado", "Resumo", "Despacho"])

        for dado in dados:
            hyperlink = Hyperlink(ref=f"A{sheet.max_row + 1}", target=url)
            sheet.cell(row=sheet.max_row + 1, column=1, value=dado["Tipo de documento"]).hyperlink = hyperlink
            sheet.cell(row=sheet.max_row, column=2, value=dado["Número do processo"])
            sheet.cell(row=sheet.max_row, column=3, value=dado["Interessado"])
            sheet.cell(row=sheet.max_row, column=4, value=dado["Resumo"])
            sheet.cell(row=sheet.max_row, column=5, value=dado["Despacho"])

        loading_label.config(text="")

        try:
            workbook.save(filename=file_name)
        except PermissionError:
            messagebox.showerror("Erro", "Feche o arquivo excel e tente novamente")
            return

        url_entry.delete(0, tk.END)

        id_url = r"/(\d+)/"
        id = re.search(id_url, url).group(1)
        messagebox.showinfo("Sucesso", f"Dados do processo {id} foram obtidos com sucesso")
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro: {str(e)}")

# Função para iniciar o processamento em segundo plano
def submit_button_clicked():
    # Crie uma thread para processar os dados
    processing_thread = threading.Thread(target=process_data)
    processing_thread.start()

app = tk.Tk()
app.title("Coletor de Dados")

# Ajusta o tamanho da janela
app.geometry("300x200")  # Largura x Altura

# Centraliza a janela na tela
app.update_idletasks()
width = app.winfo_width()
height = app.winfo_height()
x = (app.winfo_screenwidth() // 2) - (width // 2)
y = (app.winfo_screenheight() // 2) - (height // 2)
app.geometry(f"{width}x{height}+{x}+{y}")

# Cria um frame central para os widgets
frame = tk.Frame(app)
frame.pack(expand=True)

# Mensagem de Carregamento
loading_label = tk.Label(app, text="", font=("Arial", 12))
loading_label.pack(pady=10)

# Cria um campo de entrada para a URL
url_label = tk.Label(frame, text="URL do Processo:")
url_label.pack(pady=5)

url_entry = tk.Entry(frame, width=40)
url_entry.pack(pady=10)

style = ttk.Style()
style.configure("Black.TButton", background="black")

# Cria um botão para enviar
submit_button = ttk.Button(frame, text="Enviar", command=submit_button_clicked, style="Black.TButton")
submit_button.pack(pady=10)

app.mainloop()